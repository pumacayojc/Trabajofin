import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# Directorio donde se encuentran los archivos CSV
directory = 'C:/Users/Usuario/PycharmProjects/entornoanalitica/Clase_series_pca_n/Clase series de tiempo/cvss'

# Obtener todos los archivos CSV en el directorio
csv_files = [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.csv')]

# Para cada archivo CSV, crear un archivo Excel y agregar un gráfico de barras
for csv_file in csv_files:
    # Importar datos del archivo CSV
    df = pd.read_csv(csv_file, sep=';')

    # Agrupar los datos por fecha y sumar la producción por día
    daily_production = df.groupby('FECHA')['Kgprod'].sum()

    # Crear un nuevo archivo Excel y agregar el gráfico de barras
    wb = Workbook()
    ws = wb.active
    chart = BarChart()

    # Agregar los datos de producción diaria al archivo Excel y crear el gráfico de barras
    for i, (date, production) in enumerate(daily_production.items()):
        ws.cell(row=i+1, column=1, value=date)
        ws.cell(row=i+1, column=2, value=production)
    chart_data = Reference(ws, min_col=2, min_row=1, max_row=len(daily_production))
    chart.add_data(chart_data)
    chart.title = 'Producción por día'
    chart.x_axis.title = 'Fecha'
    chart.y_axis.title = 'Producción (Kg)'
    ws.add_chart(chart, 'D2')

    # Guardar el archivo Excel con el mismo nombre que el archivo CSV pero con la extensión .xlsx
    wb.save(os.path.splitext(csv_file)[0] + '.xlsx')